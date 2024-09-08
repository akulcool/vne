import subprocess
import pickle
import os
from openpyxl import load_workbook, Workbook
import json
import sys
import time

def execute_substrate_network(args_file, ch):
    with open(args_file, 'r') as file:
        arguments = file.readline().strip().split()
    arguments.append(str(ch))
    arguments.append("SN/SN.topo.pickle")
    print("\nExecuting Substrate Network with arguments:", arguments)
    mininet_script_path = '/media/sdn/New Volume/PyCharm Projects - ubuntu/Framework-3.7/mininet/Mininet.py'
    env = os.environ.copy()
    env[
        'PYTHONPATH'] = "/media/sdn/New Volume/PyCharm Projects - ubuntu/Framework-3.7/.venv/lib/python3.7/site-packages"
    python_exec = "/media/sdn/New Volume/PyCharm Projects - ubuntu/Framework-3.7/.venv/bin/python"
    command = ['sudo', '-S', sys.executable, mininet_script_path] + arguments
    print("Running command:", " ".join(command))
    subprocess.run(command, env=env)
    print("\nSubstrate Network Execution Completed.\n")

def execute_vnr_generator(args_file, vnr_gen_ch, sn_vm_gen_ch, vnr_file, sz):
    with open(args_file, 'r') as file:
        arguments = file.readline().strip().split()
    python_exec = "/media/sdn/New Volume/PyCharm Projects - ubuntu/Framework-3.7/.venv/bin/python"
    print("Executing VNR with arguments", arguments)
    subprocess.run(
        [python_exec, "mininet/VNE.generator.py", "SN/SN.topo.pickle", vnr_file, sz] + arguments + [vnr_gen_ch])
    print("VNR execution completed.")

def load_network_data(path):
    with open(path, 'rb') as f:
        data = pickle.load(f)
    return data

def calculate_total_bandwidth(topology_data):
    return sum(link['assigned_bandwidth'] for link in topology_data['links_details'])

def process_topology_data(topology_data):
    link_flags = {}
    node_flags = {}
    for link in topology_data['links_details']:
        node1, node2 = link['node1'], link['node2']
        link_flags[(node1, node2)] = False
    for node_name in topology_data:
        if node_name.startswith('h'):
            node_flags[node_name] = False
    return link_flags, node_flags

def print_vnr_details(vnr_data, vnr_id=None):
    for vnr in vnr_data:
        if vnr_id is not None and vnr['vnr_id'] != vnr_id:
            continue
        print(f"VNR{vnr['vnr_id'] + 1} Details:")
        print("VM's details:")
        for i, cpu_cores in enumerate(vnr['vm_cpu_cores'], start=1):
            print(f"VM{i} - CPU Demand: {cpu_cores}")
        print("\nVirtual Links details:")
        for i, vm_links in enumerate(vnr['vm_links'], start=1):
            if isinstance(vm_links[0], int):
                vm_links = [(vm_links[j], vm_links[j + 1]) for j in range(0, len(vm_links), 2)]
            for link in vm_links:
                vm1, vm2 = link
                print(f"VM{vm1 + 1} <--> VM{vm2 + 1}, Bandwidth demand: {vnr['bandwidth_values'][i - 1]}")

def extract_vm_to_host(data):
    print("Data received in extract_vm_to_host:", data)  # Debugging line
    if isinstance(data, list):
        vm_to_host = data[0]
        return {vm: host for vm, host in vm_to_host}
    else:
        raise ValueError("Expected data to be a list, but got something else.")

def extract_connections(data):
    connections = data[1]
    return [(conn[0], conn[1], conn[2]) if len(conn) == 3 else (conn[0], conn[1], 0) for conn in connections]

def deduct_allocated_cores(data, deductions):
    try:
        print("\nSubstrate Network CPU Available Details Before Deduction:")
        for node_name, info in data.items():
            if node_name.startswith('h'):
                print(f"{node_name}: {info['allocated_cores']}")
        for node_name, deduction_amount in deductions:
            if node_name in data:
                data[node_name]['allocated_cores'] -= deduction_amount
                data[node_name]['allocated_cores'] = max(0, data[node_name]['allocated_cores'])
        print("\nSubstrate Network CPU Available Details After Deduction:")
        for node_name, info in data.items():
            if node_name.startswith('h'):
                print(f"{node_name}: {info['allocated_cores']}")
        print("\nNode Embedding is successful.")
        return data, True
    except Exception as e:
        print(f"Error occurred during deduction: {str(e)}")
        return data, False

def deduct_allocated_bandwidth(data, connections):
    try:
        print("\nSubstrate Network Bandwidth Available Details Before Deduction:")
        for link in data['links_details']:
            print(f"{link['node1']} <-> {link['node2']} BW: {link['assigned_bandwidth']}")

        for vms, path, bandwidth in connections:
            for i in range(len(path) - 1):
                u = path[i]
                v = path[i + 1]
                for link in data['links_details']:
                    if (link['node1'] == u and link['node2'] == v) or (link['node1'] == v and link['node2'] == u):
                        print(f"Before deduction: {link['node1']} <-> {link['node2']} BW: {link['assigned_bandwidth']}")
                        link['assigned_bandwidth'] -= bandwidth
                        link['assigned_bandwidth'] = max(0, link['assigned_bandwidth'])
                        print(f"After deduction: {link['node1']} <-> {link['node2']} BW: {link['assigned_bandwidth']}")

        print("\nSubstrate Network Bandwidth Available Details After Deduction:")
        for link in data['links_details']:
            print(f"{link['node1']} <-> {link['node2']} BW: {link['assigned_bandwidth']}")
        print("\nLink Embedding is successful.")
        return data, True
    except Exception as e:
        print(f"Error occurred during bandwidth deduction: {str(e)}")
        return data, False

def append_data_to_excel(excel_file_path, data, name):
    if os.path.exists(excel_file_path):
        book = load_workbook(excel_file_path)
    else:
        book = Workbook()
    sheet = book.active
    if sheet.max_row == 1:
        headers = ["S.No", "Algorithm"] + list(data.keys())
        sheet.append(headers)
    row = [sheet.max_row + 1, name] + list(data.values())
    sheet.append(row)
    book.save(excel_file_path)

def algo(substrate_pickle_file_path, algo, vnr_pickle_file_path, excel_file_path, name):
    start_time = time.time()  # Start the timer

    SN_data = load_network_data(substrate_pickle_file_path)
    vnr_data = load_network_data(vnr_pickle_file_path)
    print(f"\n{vnr_pickle_file_path}")
    link_flags, node_flags = process_topology_data(SN_data)
    vnr_count = 0
    vm_count = 0
    s_vnr_count = 0
    Revenue = []
    cost = []
    total_vms_used_for_embedding = 0
    total_vls_used_for_embedding = 0
    num_vnrs = len(vnr_data)
    total_virtual_links = 0  # Initialize total virtual links
    total_path_length = 0  # Initialize total path length
    total_paths = 0  # Initialize total number of paths

    # Calculate initial total available CPU and BW
    initial_total_cpu = sum(info['allocated_cores'] for node_name, info in SN_data.items() if node_name.startswith('h'))
    initial_total_bw = calculate_total_bandwidth(SN_data)
    print(f"Initial Total BW: {initial_total_bw}")

    used_servers = set()  # Track used servers
    vm_to_server_mapping = {}  # Track VM to server mapping
    idle_servers = {node_name for node_name in SN_data if node_name.startswith('h')}  # Initialize idle servers

    print("Number of VNRs present:", num_vnrs)
    for idx, vnr in enumerate(vnr_data, start=1):
        vm_count += len(vnr['vm_cpu_cores'])
        total_virtual_links += len(vnr['vm_links'])  # Count virtual links
        Revenue.append(0)
        cost.append(0)
        vnr_count += 1
        backup_data = SN_data
        print(f"\nEmbedding VNR{idx}:")
        print_vnr_details(vnr_data, idx - 1)
        vnr_info = {f'VNR{idx}': {f'VM{i}': {'cpu': cpu} for i, cpu in enumerate(vnr['vm_cpu_cores'], start=1)}}
        arg = []
        arg.append(json.dumps(vnr_info))
        arg.append(json.dumps(SN_data))
        arg.append(str(idx))
        arg.append(json.dumps(vnr))
        subprocess.run(["python3", algo] + arg)
        with open('Node & Link Embedding Details.pickle', 'rb') as f:
            t = pickle.load(f)

        print("Loaded embedding details:", t)  # Debugging line

        if t is None or len(t) < 4 or t[3] == False:  # Check for embedding success
            continue
        else:
            s_vnr_count += 1
            reve = sum(vnr['vm_cpu_cores']) + sum(vnr['bandwidth_values'])
            cos = sum(vnr['vm_cpu_cores'])

            # Get VM to Host Mappings
            vm_to_host_mappings = extract_vm_to_host(t)
            print(f"VM to Host Mappings for VNR{idx}:", vm_to_host_mappings)  # Debug print
            deduction = []
            for i, cpu_cores in enumerate(vnr['vm_cpu_cores'], start=1):
                vm_key = f"VM{i}"
                if vm_key not in vm_to_host_mappings:
                    print(f"Error: {vm_key} not found in vm_to_host_mappings")  # Debug print
                    continue  # Skip this VM if not found
                server = vm_to_host_mappings[vm_key]
                used_servers.add(server)
                idle_servers.discard(server)  # Remove server from idle set if it's used
                vm_to_server_mapping[f"VNR{idx}-->{vm_key}"] = server  # Track VNR number and VM-to-server mapping

                node_flags[server] = True
                cos += cpu_cores
                deduction.append((server, cpu_cores))
            SN_data, deduction_successful = deduct_allocated_cores(SN_data, deduction)
            if deduction_successful:
                print(
                    f"CPU demand deducted from Substrate Network available CPU, after successful Node mapping of VNR{idx}.")
            else:
                print(f"Failed to deduct CPU demand from Substrate Network available CPU after Node mapping VNR{idx}.")

            # Get Connection Details
            connection_details = extract_connections(t)
            SN_data, bandwidth_deduction_successful = deduct_allocated_bandwidth(SN_data, connection_details)
            if bandwidth_deduction_successful:
                print(f"BW demand deducted from Substrate Network available BW, after successful Link mapping of VNR{idx}.")
            else:
                print(f"Failed to deduct BW demand from Substrate Network available BW after Link mapping VNR{idx}.")

            for vms, path, bandwidth in connection_details:
                if isinstance(path, int):  # Handle paths that are not lists
                    continue
                reve += bandwidth
                l = len(path)
                cos += l * bandwidth
                total_path_length += l  # Accumulate path lengths
                total_paths += 1  # Count the path

                for j in range(len(path) - 1):
                    node1 = path[j]
                    node2 = path[j + 1]
                    if (node1, node2) in link_flags:
                        link_flags[(node1, node2)] = True
                    if (node2, node1) in link_flags:
                        link_flags[(node2, node1)] = True

            Revenue[idx - 1] = reve
            cost[idx - 1] = cos
            total_vms_used_for_embedding += len(vnr['vm_cpu_cores'])
            total_vls_used_for_embedding += len(vnr['vm_links'])

    # Load the final total bandwidth from the pickle file
    with open('Node & Link Embedding Details.pickle', 'rb') as f:
        embedding_data = pickle.load(f)

        # Ensure the embedding data has the required length before accessing the final bandwidth value
        if len(embedding_data) >= 7:
            initial_total_bw = embedding_data[5]
            final_total_bw = embedding_data[6]
        else:
            initial_total_bw = initial_total_bw  # Default to initial value if not provided
            final_total_bw = calculate_total_bandwidth(SN_data)  # Default to calculated final value

    M_SN_data = load_network_data(substrate_pickle_file_path)
    O_SN_host = {}
    energy_consu = {}
    P_idle = 150
    P_full = 300
    for host, info in M_SN_data.items():
        if host.startswith('h'):
            O_SN_host[host] = info['allocated_cores']
    UP_SN_host = {}
    for host, info in SN_data.items():
        if host.startswith('h'):
            UP_SN_host[host] = info['allocated_cores']
            energy_consu[host] = round(
                P_idle + (P_full - P_idle) * ((O_SN_host[host] - UP_SN_host[host]) / O_SN_host[host]), 2)

    AR = round((s_vnr_count / vnr_count) * 100, 2)  # Acceptance Ratio as percentage
    TS = len(node_flags)
    SU = sum(1 for flag in node_flags.values() if flag is True)
    PL = len(link_flags)
    LU = sum(1 for flag in link_flags.values() if flag is True)
    NS = round((total_vms_used_for_embedding / TS), 2) if TS > 0 else 0
    ANS = round((total_vms_used_for_embedding / SU), 2) if SU > 0 else 0
    LS = round((total_vls_used_for_embedding / PL), 2) if PL > 0 else 0
    ALS = round((total_vls_used_for_embedding / LU), 2) if LU > 0 else 0
    print('\n\033[1m\033[4m' + "Performance Matrices calculations" + '\033[0m')

    print(f"Acceptance Ratio: {AR}% (Out of {vnr_count} VNRs {s_vnr_count} VNRs are accepted)")
    print(f"Total Servers: {TS}, Servers Used: {SU}, Idle Servers: {TS - SU}")
    print(f"Total Physical Links: {PL}, Links Used: {LU}, Idle Links: {PL - LU}")
    print(f"Total Virtual Links: {total_virtual_links}")  # Print total virtual links
    print(f"Nodes Stress: {NS}, Link Stress: {LS}")
    print(f"Active Nodes Stress: {ANS}, Active Link Stress: {ALS}")

    # Calculate total available CPU and BW after embedding
    after_embedding_total_cpu = sum(
        info['allocated_cores'] for node_name, info in SN_data.items() if node_name.startswith('h'))

    # Calculate total energy of used SN
    total_energy_used_sn = round(sum(energy_consu[host] for host, used in node_flags.items() if used), 2)

    # Calculate average path length
    average_path_length = round(total_path_length / total_paths, 2) if total_paths > 0 else 0

    data = {
        "VNR Pickle File Name": vnr_pickle_file_path,
        "Number of VNRs": vnr_count,
        "Acceptance Ratio": f"{AR}%",
        "Total Available Servers": TS,
        "Number of Servers Used": SU,
        "Names of Servers Used": ", ".join(sorted(used_servers)),
        "Number of Idle Servers": TS - SU,
        "Names of Idle Servers": ", ".join(sorted(idle_servers)),
        "VM-to-Server Mapping": ", ".join([f"{vm} -> {server}" for vm, server in vm_to_server_mapping.items()]),
        "Total Available Physical Links": PL,
        "Number of Links Used": LU,
        "Number of Idle Links": PL - LU,
        "Total Number of VMs": vm_count,
        "Total VMs embedded": total_vms_used_for_embedding,    # Add total VMs used for embedding
        "Total Number of VLs": total_virtual_links,
        "Total VLs embedded": total_vls_used_for_embedding,  # Add total VLs used for embedding
        "Nodes Stress": NS,
        "Active Nodes Stress": ANS,
        "Link Stress": LS,
        "Active Link Stress": ALS,
        "Before Embedding Total Available CPU of SN": initial_total_cpu,
        "After Embedding Total Available CPU of SN": after_embedding_total_cpu,
        "Before Embedding Total Available BW of SN": initial_total_bw,
        "After Embedding Total Available BW of SN": final_total_bw,
        "Average Path Length": average_path_length,
        "Total Energy of Embedded servers only in SN": f"{total_energy_used_sn} Watts"
    }

    # Calculate Avg R/C Ratio
    total_ratio = 0
    for i in range(len(Revenue)):
        if Revenue[i] != 0 and cost[i] != 0:
            total_ratio += Revenue[i] / cost[i]
    avg_rc_ratio = round(total_ratio / s_vnr_count, 2) if s_vnr_count > 0 else "N/A"
    data["Avg R/C Ratio"] = avg_rc_ratio

    Total_energy = round(sum(energy_consu[host] for host in energy_consu), 2)
    data["Total Energy of SN"] = f"{Total_energy} Watts"

    # Calculate total execution time
    total_execution_time = round(time.time() - start_time, 2)
    data["Total Execution Time"] = f"{total_execution_time} seconds"

    append_data_to_excel(excel_file_path, data, name)

    print("\nPerformance matrix appended to Excel sheet successfully.")


def main():
    excel_file_path = 'OUTPUT/parameters_output.xlsx'

    print("Enter your choice for SN and VM Distribution:")
    print("1. Random Distribution")
    print("2. Uniform Distribution")
    print("3. Normal Distribution")
    print("4. Poison Distribution")
    sn_vm_gen_ch = input()

    args_file = 'SN-Input-File.txt'
    execute_substrate_network(args_file, sn_vm_gen_ch)
    substrate_pickle_file_path = 'SN/SN.topo.pickle'

    VNRs = []

    while True:
        print("Enter your choice for VNR Distribution:")
        print("1. Random Distribution")
        print("2. Uniform Distribution")
        print("3. Normal Distribution")
        print("4. Poison Distribution")
        vnr_gen_ch = input()

        print("Enter your choice for Number of VNRs:")
        print("1. 20 VNRs")
        print("2. 40 VNRs")
        print("3. 60 VNRs")
        print("4. 80 VNRs")
        print("5. 100 VNRs")
        print("0. Exit")
        vnr_ch = int(input())

        if vnr_ch == 0:
            break

        vnr_args_file = "VNE-Input-File.txt"
        vnr_file = ""
        for i in range(10):
            if vnr_ch == 1:
                vnr_file = "VNR/vnr20." + str(i) + ".topo.pickle"
                print(vnr_file)
                execute_vnr_generator(vnr_args_file, vnr_gen_ch, sn_vm_gen_ch, vnr_file, '20')
                VNRs.append(vnr_file)
            elif vnr_ch == 2:
                vnr_file = "VNR/vnr40." + str(i) + ".topo.pickle"
                print(vnr_file)
                execute_vnr_generator(vnr_args_file, vnr_gen_ch, sn_vm_gen_ch, vnr_file, '40')
                VNRs.append(vnr_file)
            elif vnr_ch == 3:
                vnr_file = "VNR/vnr60." + str(i) + ".topo.pickle"
                print(vnr_file)
                execute_vnr_generator(vnr_args_file, vnr_gen_ch, sn_vm_gen_ch, vnr_file, '60')
                VNRs.append(vnr_file)
            elif vnr_ch == 4:
                vnr_file = "VNR/vnr80." + str(i) + ".topo.pickle"
                print(vnr_file)
                execute_vnr_generator(vnr_args_file, vnr_gen_ch, sn_vm_gen_ch, vnr_file, '80')
                VNRs.append(vnr_file)
            elif vnr_ch == 5:
                vnr_file = "VNR/vnr100." + str(i) + ".topo.pickle"
                print(vnr_file)
                execute_vnr_generator(vnr_args_file, vnr_gen_ch, sn_vm_gen_ch, vnr_file, '100')
                VNRs.append(vnr_file)

        while True:
            print("Select the Algorithm to execute:")
            print("1. CEVNE Algorithm")
            print("2. DROI Algorithm")
            print("3. Obj1+AHP")
            print("4. SCA-R")
            print("5. First Fit")
            print("6. Energy - Math")
            print("0. Go back to Number of VNRs")

            algo_ch = int(input())

            if algo_ch == 0:
                break
            elif algo_ch == 1:
                for vnr in VNRs:
                    algo(substrate_pickle_file_path, "CEVNE.py", vnr, excel_file_path, "CEVNE")
            elif algo_ch == 2:
                for vnr in VNRs:
                    algo(substrate_pickle_file_path, "DROI.py", vnr, excel_file_path, "DROI")
            elif algo_ch == 3:
                for vnr in VNRs:
                    algo(substrate_pickle_file_path, "Obj1+AHP.py", vnr, excel_file_path, "Obj1+AHP.py")
            elif algo_ch == 4:
                for vnr in VNRs:
                    algo(substrate_pickle_file_path, "SCA-R.py", vnr, excel_file_path, "SCA-R.py")
            elif algo_ch == 5:
                for vnr in VNRs:
                    algo(substrate_pickle_file_path, "First_Fit.py", vnr, excel_file_path, "First_Fit")
            elif algo_ch == 6:
                for vnr in VNRs:
                    algo(substrate_pickle_file_path, "Energy_Math.py", vnr, excel_file_path, "Energy_Math")
            else:
                print("Invalid choice, please select a valid algorithm.")

if __name__ == "__main__":
    main()